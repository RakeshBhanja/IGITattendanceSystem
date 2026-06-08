# =====================================================
# FACE RECOGNITION ATTENDANCE SYSTEM (HIGH ACCURACY)
# =====================================================

import cv2
import numpy as np
import face_recognition as fr
import os
from tkinter import *
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime

# ---------------- GLOBAL VARIABLES ----------------
known_face_encodings = []
known_face_names = []
video_capture = None
running = False
master_excel_path = None
image_folder_path = None
wb = None
ws = None
today_col = None


# ---------------- SELECT IMAGE FOLDER ----------------
def select_image_folder():
    global image_folder_path
    image_folder_path = filedialog.askdirectory()
    if image_folder_path:
        messagebox.showinfo("Success", "Image folder selected!")


# ---------------- SELECT MASTER EXCEL ----------------
def select_master_excel():
    global master_excel_path
    master_excel_path = filedialog.askopenfilename(
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if master_excel_path:
        messagebox.showinfo("Success", "Master Excel selected!")


# ---------------- LOAD FACES (MULTIPLE IMAGES SUPPORT) ----------------
def load_faces():
    global known_face_encodings, known_face_names

    known_face_encodings.clear()
    known_face_names.clear()

    for filename in os.listdir(image_folder_path):
        if filename.lower().endswith((".jpg", ".jpeg", ".png")):

            # Extract roll number before "_"
            roll_no = filename.split("_")[0].strip()

            image_path = os.path.join(image_folder_path, filename)
            image = fr.load_image_file(image_path)
            encodings = fr.face_encodings(image)

            if encodings:
                known_face_encodings.append(encodings[0])
                known_face_names.append(roll_no)

    print("Total face samples loaded:", len(known_face_encodings))


# ---------------- LOAD EXCEL ----------------
def load_excel_once():
    global wb, ws, today_col

    wb = load_workbook(master_excel_path)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")

    today_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == today:
            today_col = col
            break

    if today_col is None:
        today_col = ws.max_column + 1
        ws.cell(row=1, column=today_col).value = today
        wb.save(master_excel_path)


# ---------------- MARK ATTENDANCE ----------------
def mark_attendance(roll_no):
    global wb, ws, today_col

    roll_no = str(roll_no).strip()

    for row in range(2, ws.max_row + 1):
        excel_roll = str(ws.cell(row=row, column=1).value).strip()

        if excel_roll == roll_no:
            if ws.cell(row=row, column=today_col).value != "✔":
                ws.cell(row=row, column=today_col).value = "✔"
                wb.save(master_excel_path)
                print("Attendance marked:", roll_no)
            return


# ---------------- START ATTENDANCE ----------------
def start_attendance():
    global video_capture, running

    if not image_folder_path or not master_excel_path:
        messagebox.showerror("Error", "Select image folder and master excel first!")
        return

    load_faces()
    load_excel_once()

    video_capture = cv2.VideoCapture(0)

    if not video_capture.isOpened():
        messagebox.showerror("Error", "Camera not found!")
        return

    running = True
    process_video()


# ---------------- STOP ATTENDANCE ----------------
def stop_attendance():
    global running, video_capture

    running = False

    if video_capture is not None:
        video_capture.release()
        video_capture = None

    cv2.destroyAllWindows()
    print("System stopped.")


# ---------------- PREVIEW CAMERA ----------------
def preview_camera():
    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        cv2.imshow("Preview - Press Q to Close", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


# ---------------- VIDEO PROCESSING ----------------
def process_video():
    global running

    if not running:
        return

    ret, frame = video_capture.read()
    if not ret:
        stop_attendance()
        return

    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # Resize slightly for balance between speed & accuracy
    small_frame = cv2.resize(rgb_frame, (0, 0), fx=0.5, fy=0.5)

    face_locations = fr.face_locations(small_frame, model="hog")
    face_encodings = fr.face_encodings(small_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):

        face_distances = fr.face_distance(known_face_encodings, face_encoding)

        if len(face_distances) == 0:
            continue

        best_match_index = np.argmin(face_distances)
        distance = face_distances[best_match_index]

        if distance < 0.6:   # You can adjust (0.55 - 0.65)
            roll_no = known_face_names[best_match_index]
            mark_attendance(roll_no)

            # Scale back up face location
            top *= 2
            right *= 2
            bottom *= 2
            left *= 2

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, roll_no,
                        (left, top - 10),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.9,
                        (0, 255, 0),
                        2)

    cv2.imshow("Attendance System (Press Q to Stop)", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        stop_attendance()
        return

    root.after(20, process_video)


# ---------------- GUI ----------------
root = Tk()
root.title("Face Recognition Attendance System")
root.geometry("420x350")

Label(root, text="Face Recognition Attendance System",
      font=("Arial", 14)).pack(pady=10)

Button(root, text="Select Image Folder",
       command=select_image_folder,
       width=30).pack(pady=5)

Button(root, text="Select Master Excel",
       command=select_master_excel,
       width=30).pack(pady=5)

Button(root, text="Preview Camera",
       command=preview_camera,
       bg="blue", fg="white",
       width=30).pack(pady=10)

Button(root, text="Start Attendance",
       command=start_attendance,
       bg="green", fg="white",
       width=30).pack(pady=5)

Button(root, text="Stop Attendance",
       command=stop_attendance,
       bg="red", fg="white",
       width=30).pack(pady=5)

root.mainloop()